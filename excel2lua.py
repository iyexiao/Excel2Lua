# -*- coding: utf-8 -*-
#Author: yexiao
#Date: 2017-07-07
#Purpose: 解析excel数据表到lua

import os
import sys
import string
import xlrd
import json
import openpyxl
reload(sys)
sys.setdefaultencoding('utf8')

_errorDes = ["数据错误：","文件名","行数","列数"]
_isError = False

global excelPath
global luaPath


global isXls

# 按类型解析数据
dataType = ["int","string","bool","table"]
# 查找所有 xls
def findAllFile(callback):
    fileList = os.listdir(excelPath)
    for f in fileList:

        # 检查是否有错误
        if _isError:
            break

        filePath = os.path.join(excelPath,f)

        if f[0] == "." or f.find(".svn") > 0 or f.find(".DS_Store") > 0 or f.startswith("~$"):
            continue
    
        if os.path.isdir(filePath):
            findAllFile(filePath,callback)
        else:
            if f.endswith(".xls") or f.endswith(".xlsx"):
                callback(filePath,f)

# 解析 excel的type类型
# ctype类型 0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
def parseTitleType(rowData,col):
    tableKey = []
    tableType=[]
    for c in range(col):
        typeValue = None
        keyValue = None
        if isXls:
            typeValue = rowData.cell(1,c).value
            keyValue = rowData.cell(2,c).value
        else:
            typeValue = rowData.cell(row=2,column=c+1).value
            keyValue = rowData.cell(row=3,column=c+1).value
        if typeValue == None:
            break
        tmp = typeValue.split('|')
        # 这里判断只导出客户端的
        if len(tmp) == 2 :
            if tmp[1]=='client':
                typeValue = tmp[0]
            else:
                typeValue = None
        tableKey.append(keyValue)
        tableType.append(typeValue)
    return tableKey,tableType

# 获取表格数据
def getDataByExcel(pTabData,pRow,tblKey):
    ret = []
    for row in range(pRow):
        coltab=[]
        for col in range(len(tblKey)):
            tmp = None
            global isXls
            if isXls:
                tmp = pTabData.cell(row+3,col).value
            else:
                tmp = pTabData.cell(row=row+4,column=col+1).value
            if tmp == None:
                tmp = 'nil'
            coltab.append(tmp)
        ret.append(coltab)
    return ret


# ******************************************************* 格式化输出数据
# 文件介绍
_fileDes='''-- Filename: %s.lua
-- Author: auto-created by yexiao`s ParseExcel(to lua) tool.
-- methods: X.getDataById(id), X.getArrDataByField(fieldName, fieldValue)
-- Function: no description.\n
'''
_fileKeys = "local keys = {\n\t\"%s\"\n}\n\n"
_fileData = "local data = {\n%s}\n\n"
_fileDataItem = "[%s]={%s},\n"
_fileExports = "cc.exports.%s = {}\n"
_fileFun ='''function %s.getDataById(id)
    if not id or type(id) ~= 'number' then return nil end
    local tmp = data[id]
    if not tmp then return nil end
    local tbl = {}
    for k,v in pairs(keys) do
         tbl[v] = tmp[k]
     end
     return tbl
end'''
# *******************************************************
# 将excel行数据整理成lua的table
def parseData(pData,pType):
    ret = ""
    # print pData
    # print pType
    for row in range(len(pData)):
        item=[]
        for col in range(len(pType)):
            tmp = pData[row][col]
            if tmp == None or tmp == '':
                tmp = 'nil'
            if pType[col] == dataType[0]:
                if tmp != 'nil':
                    tmp = int(tmp)
            elif pType[col] == dataType[1]:
                if tmp != 'nil':
                    tmp= "\"%s\"" % (tmp.replace('\"','\\\"'))
            elif pType[col] == dataType[2]:
                if tmp == 'false' or tmp == 'nil':
                    tmp = 'false'
            elif pType[col] == dataType[3]:
                if tmp != 'nil':
                    rrr = []
                    vtmp = tmp.split(',')
                    if len(vtmp)>0:
                        for i in range(len(vtmp)):
                            vvtmp = vtmp[i].split('=')
                            if len(vvtmp) == 2:
                                fff = vvtmp[0]
                                sss = ("\"%s\"" % (vvtmp[1].replace('\"','\\\"')))
                                if sss == '"true"':
                                    sss = 'true'
                                elif sss == '"false"':
                                    sss = 'false'
                                if fff.isdigit() :
                                    rrr.append('['+str(fff)+']='+sss)
                                else:
                                    rrr.append(str(fff)+'='+sss)
                            else:
                                rrr.append(str(vtmp[i]))
                        tmp = '{%s}' %','.join(rrr)
                    else:
                        tmp = '{%s}' % (tmp.replace('\"','\\\"'))
            item.append(str(tmp))
            # print item
        fileItem = _fileDataItem%(row+1,','.join(item))
        ret = ret+fileItem
    return ret

# 构造输出lua
def getLuaText(title,pData,fileName,pType,pKey):

    # 文件介绍
    luaFileName = "DB_%s" % (title.capitalize())
    fileStr1 = _fileDes%(luaFileName)
    # 文件keys
    fileStr2 = _fileKeys%('\",\"'.join(pKey))

    # 文件数据内容
    fileStr3 = _fileData%(parseData(pData,pType))
    # cocos 全局变量
    fileStr4 = _fileExports%(luaFileName)
    # 文件返回
    fileStr5 = _fileFun%(luaFileName)
    fileStr = fileStr1+fileStr2+fileStr3+fileStr4+fileStr5

    return luaFileName,fileStr
# 写文件
def writeFile(filePath,fileData):
    f = open(filePath,"w")
    f.write(fileData)
    f.close()
# 写文件
def writeToLua(filePath,fileName,fileData):
    global excelPath
    global luaPath
    fDirPath = os.path.dirname(filePath)
    fDirPath = fDirPath.replace(excelPath,luaPath)
    if not os.path.exists(fDirPath):
        os.mkdir(fDirPath)
    filePath = os.path.join(fDirPath,"%s.lua" % (fileName))
    writeFile(filePath,fileData)
#拆分excel
def parseExcel(filePath,fileName):
    global isXls
    if fileName.endswith(".xls"):
        isXls = True
        # 读取数据
        excel = xlrd.open_workbook(filePath)
        #获取workbook中所有的表格  
        sheets = excel.sheet_names() 
        #循环遍历所有sheet  
        for i in range(len(sheets)):
            sheet = excel.sheet_names()[i]
            pTabData = excel.sheet_by_index(i)
            if pTabData.nrows > 3 and pTabData.ncols > 0 :
                tabTitle = parseTitleType(pTabData,pTabData.ncols)
                data =getDataByExcel(pTabData,pTabData.nrows - 3,tabTitle[0])
                title = sheet.split('|')
                if len(title) == 2:
                    # 解析后的数据[文件名，文件内容]
                    filedata = getLuaText(title[1],data,fileName,tabTitle[1],tabTitle[0])
                    # 写入文件
                    writeToLua(filePath,filedata[0],filedata[1])
                else:
                    print 'excel sheet配置错误'
                    sys.exit(1)
    else:
        isXls = False
        excel = openpyxl.load_workbook(filePath)
        sheets = excel.get_sheet_names() 
        #循环遍历所有sheet  
        for i in range(len(sheets)):  
            sheet= excel.get_sheet_by_name(sheets[i])
            if sheet.max_row > 3 and sheet.max_column > 0 :
                tabTitle = parseTitleType(sheet,sheet.max_column)
                data = getDataByExcel(sheet,sheet.max_row - 3,tabTitle[0])
                title = sheet.title.split('|')
                if len(title) == 2:
                    # 解析后的数据[文件名，文件内容]
                    filedata = getLuaText(title[1],data,fileName,tabTitle[1],tabTitle[0])
                    # 写入文件
                    writeToLua(filePath,filedata[0],filedata[1])
                else:
                    print 'excel sheet配置错误'
                    sys.exit(1)



# 1.excel 文件夹路径 2.lua文件路径
if __name__ == "__main__":
    global excelPath
    global luaPath
    if len(sys.argv) <= 3:
        excelPath = sys.argv[1]
        if not os.path.exists(excelPath):
            print 'excelPath 路径不存在'
            sys.exit(1)
        luaPath = sys.argv[2]
        # 检查并创建目录
        if not os.path.exists(luaPath):
            os.makedirs(luaPath)

        findAllFile(parseExcel)
        print 'parse excel success!'
    else:
        print '参数不正确'
        sys.exit(1)
